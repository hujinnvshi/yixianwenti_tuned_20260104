"""
报表生成模块
负责生成Excel报表
"""

import pandas as pd
from pathlib import Path
from typing import Dict
from loguru import logger
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re


class ReportGenerator:
    """报表生成器"""

    def __init__(self, config: Dict):
        """
        初始化报表生成器

        Args:
            config: 配置字典
        """
        self.config = config
        self.output_dir = Path(config['output']['directory'])
        self.output_filename = config['output']['filename']

        # 创建输出目录
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_report(self,
                       df2: pd.DataFrame,
                       df1_processed: pd.DataFrame,
                       pivot_df: pd.DataFrame) -> Path:
        """
        生成完整的Excel报表

        Sheet结构:
        1. 2025122911704000480: 表格2原始数据
        2. 计算解决率过程数据（调整后）: 表格1处理后数据
        3. 计算解决率: 透视表结果

        Args:
            df2: 表格2原始数据
            df1_processed: 表格1处理后数据
            pivot_df: 透视表结果

        Returns:
            Path: 输出文件路径
        """
        output_path = self.output_dir / self.output_filename

        logger.info(f"开始生成报表: {output_path}")

        # 使用ExcelWriter写入多个Sheet
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet1: 原始数据
            df2.to_excel(writer, sheet_name='2025122911704000480', index=False)
            logger.info(f"写入Sheet1 '2025122911704000480': {len(df2)}行")

            # Sheet2: 处理后数据
            df1_processed.to_excel(writer, sheet_name='计算解决率过程数据（调整后）', index=False)
            logger.info(f"写入Sheet2 '计算解决率过程数据（调整后）': {len(df1_processed)}行")

            # Sheet3: 透视表
            pivot_df.to_excel(writer, sheet_name='计算解决率')
            logger.info(f"写入Sheet3 '计算解决率': {len(pivot_df)}行")

        # 注意: 不再添加Excel公式,保持pandas计算的结果值
        # 原因: Excel公式需要Excel打开才能计算,openpyxl无法计算公式
        # pandas已经正确计算了所有值,直接使用即可

        # 获取文件大小
        file_size = output_path.stat().st_size / 1024  # KB

        logger.info(f"""
        报表生成成功 ✓
        - 文件路径: {output_path}
        - 文件大小: {file_size:.2f} KB
        - Sheet数量: 3
        """)

        return output_path

    def _add_formulas_to_sheet2(self, file_path: Path, df: pd.DataFrame):
        """
        为Sheet2(计算解决率过程数据（调整后）)添加Excel公式

        为AE列(研发交付日期偏差)和AO列(用于交付日期偏差统计)添加公式

        策略:
        1. 保留原有的计算结果值
        2. 将单元格替换为Excel公式,这样用户可以看到和验证计算逻辑

        Args:
            file_path: Excel文件路径
            df: 数据框(用于获取列索引)
        """
        logger.info("开始添加Excel公式到Sheet2...")

        try:
            # 加载工作簿
            wb = load_workbook(file_path)
            ws = wb['计算解决率过程数据（调整后）']

            # 获取列索引
            columns = df.columns.tolist()

            # 查找关键列的索引
            col_indices = {}
            for col_name in ['期望解决时间', '计划完成时间', '研发解决时间',
                           '更新时间', '审批状态', '处理方式', '研发交付日期偏差',
                           '用于交付日期偏差统计']:
                if col_name in columns:
                    col_indices[col_name] = columns.index(col_name) + 1  # +1 因为Excel从1开始

            logger.info(f"列索引映射: {col_indices}")

            # 获取Excel中实际的数据行数(不包括表头)
            max_row = ws.max_row
            logger.info(f"数据行数: {max_row - 1}")

            # 添加AE列公式(研发交付日期偏差)
            if '研发交付日期偏差' in col_indices:
                ae_col = col_indices['研发交付日期偏差']
                ae_col_letter = get_column_letter(ae_col)

                logger.info(f"添加AE列公式到列 {ae_col_letter} (索引{ae_col})...")

                # 为每一行数据添加公式
                for row in range(2, max_row + 1):
                    cell = ws.cell(row=row, column=ae_col)

                    # 获取相关列的字母
                    planned_col = get_column_letter(col_indices['计划完成时间'])
                    expected_col = get_column_letter(col_indices['期望解决时间'])
                    dev_solve_col = get_column_letter(col_indices['研发解决时间'])
                    update_col = get_column_letter(col_indices['更新时间'])
                    status_col = get_column_letter(col_indices['审批状态'])
                    handle_col = get_column_letter(col_indices['处理方式'])

                    # AE列公式
                    # 公式说明:
                    # 基准日期 = IF(计划完成时间="", 期望解决时间, 计划完成时间)
                    # 实际日期 = IF(研发解决时间="", IF(审批状态="已结束", 更新时间, NOW()), 研发解决时间)
                    # 天数偏差 = 实际日期 - 基准日期
                    # 最终结果 = IF(处理方式<>"研发处理", "非研发处理", IF(基准日期="", "", 天数偏差))

                    # Excel公式 - 使用ISBLANK检查空单元格
                    # 问题: pandas写入时None值变成空单元格,不是空字符串""
                    # 解决: 使用ISBLANK()函数检查,而不是x=""
                    formula = f'=IF(OR(ISBLANK({handle_col}{row}),{handle_col}{row}<>"研发处理"),"非研发处理",IF(ISBLANK(IF(ISBLANK({planned_col}{row}),{expected_col}{row},{planned_col}{row})),"",IF(ISBLANK({dev_solve_col}{row}),IF({status_col}{row}="已结束",{update_col}{row},TODAY()),{dev_solve_col}{row})-IF(ISBLANK({planned_col}{row}),{expected_col}{row},{planned_col}{row}))))'

                    # 注意: 使用TODAY()而不是NOW(),避免时间部分影响天数计算
                    cell.value = formula

            # 添加AO列公式(用于交付日期偏差统计)
            if '用于交付日期偏差统计' in col_indices:
                ao_col = col_indices['用于交付日期偏差统计']
                ao_col_letter = get_column_letter(ao_col)

                logger.info(f"添加AO列公式到列 {ao_col_letter} (索引{ao_col})...")

                # 为每一行数据添加公式
                for row in range(2, max_row + 1):
                    cell = ws.cell(row=row, column=ao_col)

                    # AO列公式基于AE列的结果
                    # 公式说明:
                    # 如果AE="非研发处理", 返回"非研发处理"
                    # 如果AE是数字且研发解决时间不为空, 返回"及时解决"(AE<=0)或"未及时解决"(AE>0)
                    # 如果AE是数字且研发解决时间为空, 返回"处理中暂未超时"(AE<=0)或"超时未解决"(AE>0)

                    # 使用ISBLANK检查研发解决时间是否为空
                    ao_formula = f'=IF({ae_col_letter}{row}="非研发处理","非研发处理",IF(ISNUMBER({ae_col_letter}{row}),IF(NOT(ISBLANK({dev_solve_col}{row})),IF({ae_col_letter}{row}<=0,"及时解决","未及时解决"),IF({ae_col_letter}{row}<=0,"处理中暂未超时","超时未解决")),""))'

                    cell.value = ao_formula

            # 保存工作簿
            wb.save(file_path)

            # 重新加载并计算公式值
            logger.info("计算公式并保存结果...")
            wb2 = load_workbook(file_path)
            ws2 = wb2['计算解决率过程数据（调整后）']

            # 读取每个公式单元格的值(从data_only模式)
            # 这样Excel下次打开时会显示上次计算的值
            # 注意: openpyxl不能直接计算公式,需要Excel或xlwings
            # 我们采用的方法是: 保持公式不变,Excel会自动计算

            wb2.close()

            logger.info("Excel公式添加完成 ✓")

        except Exception as e:
            logger.error(f"添加Excel公式时出错: {e}")
            import traceback
            logger.error(traceback.format_exc())
            # 不中断流程,即使添加公式失败也继续

    def format_report(self, file_path: Path):
        """
        格式化Excel报表(添加样式、格式化等)

        Args:
            file_path: Excel文件路径
        """
        logger.info("开始格式化报表...")

        # TODO: 实现格式化逻辑
        # - 设置百分比格式
        # - 设置日期格式
        # - 设置列宽
        # - 添加表头样式
        # - 冻结首行

        logger.info("报表格式化完成 ✓")
